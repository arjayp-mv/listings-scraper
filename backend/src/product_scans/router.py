# =============================================================================
# Product Scans Domain - API Router
# =============================================================================
# Purpose: REST API endpoints for product scan jobs
# Public API: router
# Dependencies: fastapi, sqlalchemy, service, schemas
# =============================================================================

from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import io
import csv

from ..database import get_db
from ..pagination import PaginationParams, get_pagination_params, create_paginated_response
from .service import ProductScanService
from .models import JobStatus, ItemStatus
from .schemas import (
    ProductScanJobCreate,
    ProductScanJobResponse,
    ProductScanJobDetailResponse,
    ProductScanJobListResponse,
    ProductScanJobResultsResponse,
    ProductScanItemWithSkuResponse,
    ProductScanSummary,
    ProductScanDashboardStats,
)
from ..channel_skus.schemas import BulkScanRequest


router = APIRouter(prefix="/api/product-scans", tags=["Product Scans"])


# ===== Job CRUD Endpoints =====


@router.post("", response_model=ProductScanJobResponse, status_code=201)
async def create_product_scan_job(
    data: ProductScanJobCreate,
    db: Session = Depends(get_db),
):
    """
    Create a new product scan job.

    Accepts a list of Channel SKU code + ASIN pairs. Creates Channel SKUs
    if they don't exist.
    """
    service = ProductScanService(db)

    listings = [
        {
            "sku_code": item.sku_code,
            "channel_sku_code": item.channel_sku_code,
            "asin": item.asin,
        }
        for item in data.listings
    ]

    job = service.create_job(
        job_name=data.job_name,
        marketplace=data.marketplace,
        listings=listings,
    )

    return ProductScanJobResponse.model_validate(job)


@router.post("/from-channel-skus", response_model=ProductScanJobResponse, status_code=201)
async def create_scan_from_channel_skus(
    data: BulkScanRequest,
    job_name: str = Query(..., min_length=1, max_length=255),
    db: Session = Depends(get_db),
):
    """
    Create a product scan job from existing Channel SKU IDs.

    Used by the "All Listings" page to queue selected items for scanning.
    """
    service = ProductScanService(db)

    try:
        job = service.create_job_from_channel_skus(
            job_name=job_name,
            marketplace=data.marketplace or "com",
            channel_sku_ids=data.channel_sku_ids,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return ProductScanJobResponse.model_validate(job)


@router.get("", response_model=ProductScanJobListResponse)
async def list_product_scan_jobs(
    status: Optional[str] = Query(None, description="Filter by status"),
    search: Optional[str] = Query(None, description="Search job name"),
    pagination: PaginationParams = Depends(get_pagination_params),
    db: Session = Depends(get_db),
):
    """List product scan jobs with pagination and filters."""
    service = ProductScanService(db)

    # Parse status filter
    status_enum = None
    if status:
        try:
            status_enum = JobStatus(status)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid status: {status}")

    jobs, total = service.list_jobs(
        offset=pagination.offset,
        limit=pagination.limit,
        status=status_enum,
        search=search,
    )

    # Add progress percentage
    items = []
    for job in jobs:
        progress = 0.0
        if job.total_listings > 0:
            progress = (
                (job.completed_listings + job.failed_listings) / job.total_listings
            ) * 100

        items.append(
            ProductScanJobDetailResponse(
                id=job.id,
                job_name=job.job_name,
                status=job.status,
                marketplace=job.marketplace,
                total_listings=job.total_listings,
                completed_listings=job.completed_listings,
                failed_listings=job.failed_listings,
                error_message=job.error_message,
                created_at=job.created_at,
                started_at=job.started_at,
                completed_at=job.completed_at,
                progress_percent=round(progress, 1),
            )
        )

    return create_paginated_response(items, total, pagination)


@router.get("/stats", response_model=ProductScanDashboardStats)
async def get_product_scan_stats(
    db: Session = Depends(get_db),
):
    """Get product scan statistics for dashboard."""
    service = ProductScanService(db)
    stats = service.get_dashboard_stats()

    return ProductScanDashboardStats(
        total_jobs=stats["total_jobs"],
        total_listings_scanned=stats["total_listings_scanned"],
        jobs_by_status=stats["jobs_by_status"],
        recent_jobs=[
            ProductScanJobResponse.model_validate(j) for j in stats["recent_jobs"]
        ],
    )


@router.get("/{job_id}", response_model=ProductScanJobDetailResponse)
async def get_product_scan_job(
    job_id: int,
    db: Session = Depends(get_db),
):
    """Get product scan job details."""
    service = ProductScanService(db)
    job = service.get_by_id(job_id)

    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    # For running jobs, get real-time progress from item statuses
    if job.status == JobStatus.RUNNING:
        completed, failed = service.get_real_time_progress(job.id)
    else:
        completed = job.completed_listings
        failed = job.failed_listings

    progress = 0.0
    if job.total_listings > 0:
        progress = ((completed + failed) / job.total_listings) * 100

    return ProductScanJobDetailResponse(
        id=job.id,
        job_name=job.job_name,
        status=job.status,
        marketplace=job.marketplace,
        total_listings=job.total_listings,
        completed_listings=completed,
        failed_listings=failed,
        error_message=job.error_message,
        created_at=job.created_at,
        started_at=job.started_at,
        completed_at=job.completed_at,
        progress_percent=round(progress, 1),
    )


@router.get("/{job_id}/results", response_model=ProductScanJobResultsResponse)
async def get_product_scan_results(
    job_id: int,
    status: Optional[str] = Query(None, description="Filter by item status"),
    pagination: PaginationParams = Depends(get_pagination_params),
    db: Session = Depends(get_db),
):
    """Get paginated results for a product scan job."""
    service = ProductScanService(db)
    job = service.get_by_id(job_id)

    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    # Parse status filter
    status_enum = None
    if status:
        try:
            status_enum = ItemStatus(status)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid status: {status}")

    items, total = service.get_job_items(
        job_id=job_id,
        offset=pagination.offset,
        limit=pagination.limit,
        status=status_enum,
    )

    # Transform items
    response_items = []
    for item in items:
        asin_changed = (
            item.scraped_asin is not None
            and item.scraped_asin != item.input_asin
        )
        response_items.append(
            ProductScanItemWithSkuResponse(
                id=item.id,
                job_id=item.job_id,
                channel_sku_id=item.channel_sku_id,
                channel_sku_code=item.channel_sku.channel_sku_code,
                input_asin=item.input_asin,
                status=item.status,
                scraped_rating=item.scraped_rating,
                scraped_review_count=item.scraped_review_count,
                scraped_title=item.scraped_title,
                scraped_asin=item.scraped_asin,
                error_message=item.error_message,
                started_at=item.started_at,
                completed_at=item.completed_at,
                asin_changed=asin_changed,
            )
        )

    # Get summary
    summary_data = service.get_job_summary(job_id)
    summary = ProductScanSummary(**summary_data)

    paginated = create_paginated_response(response_items, total, pagination)
    paginated["summary"] = summary

    return ProductScanJobResultsResponse(**paginated)


# ===== Export Endpoints =====


@router.get("/{job_id}/export/csv")
async def export_scan_results_csv(
    job_id: int,
    db: Session = Depends(get_db),
):
    """Export scan results as CSV."""
    service = ProductScanService(db)
    job = service.get_by_id(job_id)

    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    items, _ = service.get_job_items(job_id=job_id, offset=0, limit=10000)

    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "Channel SKU",
        "Input ASIN",
        "Status",
        "Rating",
        "Review Count",
        "Product Title",
        "Scraped ASIN",
        "ASIN Changed",
        "Error",
    ])

    for item in items:
        asin_changed = (
            item.scraped_asin and item.scraped_asin != item.input_asin
        )
        writer.writerow([
            item.channel_sku.channel_sku_code,
            item.input_asin,
            item.status.value,
            str(item.scraped_rating) if item.scraped_rating else "",
            str(item.scraped_review_count) if item.scraped_review_count else "",
            item.scraped_title or "",
            item.scraped_asin or "",
            "Yes" if asin_changed else "No",
            item.error_message or "",
        ])

    output.seek(0)
    filename = f"product_scan_{job_id}_results.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{job_id}/export/excel")
async def export_scan_results_excel(
    job_id: int,
    db: Session = Depends(get_db),
):
    """Export scan results as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    service = ProductScanService(db)
    job = service.get_by_id(job_id)

    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    items, _ = service.get_job_items(job_id=job_id, offset=0, limit=10000)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Scan Results"

    # Header styling
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    headers = [
        "Channel SKU",
        "Input ASIN",
        "Status",
        "Rating",
        "Review Count",
        "Product Title",
        "Scraped ASIN",
        "ASIN Changed",
        "Error",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill

    # Data rows
    for row, item in enumerate(items, 2):
        asin_changed = (
            item.scraped_asin and item.scraped_asin != item.input_asin
        )
        ws.cell(row=row, column=1, value=item.channel_sku.channel_sku_code)
        ws.cell(row=row, column=2, value=item.input_asin)
        ws.cell(row=row, column=3, value=item.status.value)
        ws.cell(row=row, column=4, value=float(item.scraped_rating) if item.scraped_rating else None)
        ws.cell(row=row, column=5, value=item.scraped_review_count)
        ws.cell(row=row, column=6, value=item.scraped_title)
        ws.cell(row=row, column=7, value=item.scraped_asin)
        ws.cell(row=row, column=8, value="Yes" if asin_changed else "No")
        ws.cell(row=row, column=9, value=item.error_message)

    # Adjust column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 30

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"product_scan_{job_id}_results.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ===== Job Actions =====


@router.post("/{job_id}/cancel", response_model=ProductScanJobResponse)
async def cancel_product_scan_job(
    job_id: int,
    db: Session = Depends(get_db),
):
    """Cancel a queued or running job."""
    service = ProductScanService(db)
    job = service.get_by_id(job_id)

    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    try:
        service.cancel_job(job)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return ProductScanJobResponse.model_validate(job)


@router.post("/{job_id}/retry-failed", response_model=ProductScanJobResponse)
async def retry_failed_items(
    job_id: int,
    db: Session = Depends(get_db),
):
    """Retry failed items in a job."""
    service = ProductScanService(db)
    job = service.get_by_id(job_id)

    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    count = service.retry_failed_items(job)

    if count == 0:
        raise HTTPException(status_code=400, detail="No failed items to retry")

    # Refresh job
    db.refresh(job)
    return ProductScanJobResponse.model_validate(job)


@router.delete("/{job_id}", status_code=204)
async def delete_product_scan_job(
    job_id: int,
    db: Session = Depends(get_db),
):
    """Delete a product scan job and all its items."""
    service = ProductScanService(db)
    job = service.get_by_id(job_id)

    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    service.delete_job(job)
